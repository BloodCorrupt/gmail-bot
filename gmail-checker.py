from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementNotInteractableException
from selenium.webdriver.chrome.options import Options
from concurrent.futures import ThreadPoolExecutor
import openpyxl
import time
from threading import Lock

# Global variables
excel_lock = Lock()
work_count = 0
max_threads = 0


def login_to_gmail(email, password, recovery_email, change_recovery, change_pass, 
                   change_fname, change_lname, row_index, workbook):
    """Login to Gmail and perform account modifications"""
    global work_count
    
    try:
        print(f'Starting login attempt for {email}...')
        print(f'Working Account Now: {max_threads + work_count}...')
        
        # Initialize Chrome driver
        chrome_options = Options()
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--incognito')
        driver = webdriver.Chrome(options=chrome_options)
        
        try:
            # Navigate to Gmail login
            login_url = 'https://accounts.google.com/v3/signin/identifier?authuser=0&continue=https%3A%2F%2Fmail.google.com%2Fmail&ec=GAlAFw&hl=en&service=mail&flowName=GlifWebSignIn&flowEntry=AddSession'
            driver.get(login_url)
            time.sleep(3)
            
            # Enter email
            try:
                email_input = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.ID, 'identifierId'))
                )
                email_input.send_keys(email)
                email_input.send_keys(Keys.RETURN)
                time.sleep(10)
            except (TimeoutException, NoSuchElementException, ElementNotInteractableException) as e:
                print(f'Failed to enter email for {email}: {e}')
                result = 'Email entry failed'
                save_result(workbook, row_index, result)
                driver.quit()
                return
            
            # Check for initial errors (account not found, etc.)
            try:
                error_message = driver.find_element(By.ID, 'headingText')
                error_text = error_message.text
                
                if 'Welcome' not in error_text and 'Sign in' not in error_text:
                    # This is an error message
                    print(f'Error detected: {error_text}')
                    result = error_text if error_text else 'Unknown error'
                    save_result(workbook, row_index, result)
                    driver.quit()
                    work_count += 1
                    return
            except (TimeoutException, NoSuchElementException):
                pass
            
            # Try to enter password
            try:
                password_input = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.NAME, 'Passwd'))
                )
                password_input.send_keys(password)
                password_input.send_keys(Keys.RETURN)
                time.sleep(10)
            except (TimeoutException, NoSuchElementException, ElementNotInteractableException) as e:
                print(f'Failed to enter password for {email}: {e}')
                result = 'Password entry failed'
                save_result(workbook, row_index, result)
                driver.quit()
                return
            
            # Handle recovery email verification if needed
            try:
                confirm_email_button = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '(//div[@class="l5PPKe"])[3]'))
                )
                confirm_email_button.click()
                time.sleep(10)
                
                recovery_email_input = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, 'knowledge-preregistered-email-response'))
                )
                recovery_email_input.send_keys(recovery_email)
                time.sleep(5)
                recovery_email_input.send_keys(Keys.RETURN)
                time.sleep(10)
            except (TimeoutException, NoSuchElementException, ElementNotInteractableException):
                print('Recovery email verification not required or failed')
            
            # Check for account issues (disabled, 2FA, etc.)
            try:
                error_message = driver.find_element(By.ID, 'headingText')
                error_text = error_message.text.lower()
                
                if 'disabled' in error_text:
                    result = 'Account disabled'
                    save_result(workbook, row_index, result)
                    driver.quit()
                    work_count += 1
                    return
                elif 'verify it\'s you' in error_text or '2-step' in error_text:
                    result = '2-step verification required'
                    save_result(workbook, row_index, result)
                    driver.quit()
                    work_count += 1
                    return
            except (TimeoutException, NoSuchElementException):
                pass
            
            # Wait for successful login (check for inbox URL)
            try:
                WebDriverWait(driver, 15).until(EC.url_contains('#inbox'))
                result = 'Login successful'
                print(result)
            except TimeoutException:
                result = 'Login verification timeout'
                save_result(workbook, row_index, result)
                driver.quit()
                return
            
            # Perform optional account modifications
            perform_account_changes(driver, workbook, row_index, change_recovery, 
                                   change_pass, change_fname, change_lname)
            
            # Final result save
            result = 'Completed'
            save_result(workbook, row_index, result)
            
        except Exception as e:
            print(f'Unexpected error during login for {email}: {e}')
            result = f'Error: {str(e)[:100]}'
            save_result(workbook, row_index, result)
        finally:
            driver.quit()
            work_count += 1
            
    except Exception as e:
        print(f'Critical error for {email}: {e}')


def perform_account_changes(driver, workbook, row_index, change_recovery, change_pass, 
                           change_fname, change_lname):
    """Perform account modifications if requested"""
    try:
        sheet = workbook.active
        row = sheet[row_index]
        
        # Change recovery email if requested
        if row[3].value:
            try:
                print('Attempting to change recovery email...')
                recovery_url = 'https://myaccount.google.com/recovery/email'
                driver.get(recovery_url)
                time.sleep(10)
                
                edit_button = driver.find_element(By.XPATH, '//button[@class="pYTkkf-Bz112c-LgbsSe wMI9H Qd9OXe"]')
                edit_button.click()
                time.sleep(5)
                
                recovery_input = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, '//input[@type="email"]'))
                )
                time.sleep(2)
                recovery_input.clear()
                time.sleep(1)
                recovery_input.send_keys(change_recovery)
                time.sleep(2)
                recovery_input.send_keys(Keys.ENTER)
                time.sleep(5)
                print('Recovery email changed successfully')
            except Exception as e:
                print(f'Failed to change recovery email: {e}')
        
        # Change password if requested
        if row[4].value:
            try:
                print('Attempting to change password...')
                password_url = 'https://myaccount.google.com/signinoptions/password'
                driver.get(password_url)
                time.sleep(10)
                
                # Try both possible password input field IDs
                for field_id in ['i6', 'i5']:
                    try:
                        pwd_input = WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.ID, field_id))
                        )
                        time.sleep(3)
                        pwd_input.send_keys(change_pass)
                        time.sleep(3)
                        break
                    except (TimeoutException, NoSuchElementException):
                        continue
                
                # Try confirm password field
                for field_id in ['i12', 'i11']:
                    try:
                        pwd_confirm = WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.ID, field_id))
                        )
                        time.sleep(3)
                        pwd_confirm.send_keys(change_pass)
                        pwd_confirm.send_keys(Keys.RETURN)
                        time.sleep(10)
                        break
                    except (TimeoutException, NoSuchElementException):
                        continue
                
                # Click OK button if it appears
                try:
                    ok_button = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'button[data-mdc-dialog-action="ok"]'))
                    )
                    if ok_button.is_displayed():
                        driver.execute_script('arguments[0].click();', ok_button)
                        time.sleep(10)
                        print('Password changed successfully')
                except (TimeoutException, NoSuchElementException):
                    print('OK button not found')
            except Exception as e:
                print(f'Failed to change password: {e}')
        
        # Change first name if requested
        if row[5].value:
            try:
                print('Attempting to change first name...')
                profile_url = 'https://myaccount.google.com/profile/name/edit'
                driver.get(profile_url)
                time.sleep(10)
                
                fname_input = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="i6"]'))
                )
                time.sleep(3)
                fname_input.clear()
                time.sleep(1)
                fname_input.send_keys(change_fname)
                time.sleep(2)
                
                save_button = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '//button[.//span[@class="UywwFc-vQzf8d"]]'))
                )
                save_button.click()
                time.sleep(10)
                print('First name changed successfully')
            except Exception as e:
                print(f'Failed to change first name: {e}')
        
        # Change last name if requested
        if row[6].value:
            try:
                print('Attempting to change last name...')
                profile_url = 'https://myaccount.google.com/profile/name/edit'
                driver.get(profile_url)
                time.sleep(10)
                
                lname_input = WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, '//input[@id="i7"]'))
                )
                time.sleep(3)
                lname_input.clear()
                time.sleep(1)
                lname_input.send_keys(change_lname)
                time.sleep(2)
                
                save_button = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.XPATH, '//button[.//span[@class="UywwFc-vQzf8d"]]'))
                )
                save_button.click()
                time.sleep(10)
                print('Last name changed successfully')
            except Exception as e:
                print(f'Failed to change last name: {e}')
    except Exception as e:
        print(f'Error in account changes: {e}')


def save_result(workbook, row_index, result):
    """Save result to Excel file"""
    try:
        with excel_lock:
            sheet = workbook.active
            sheet.cell(row=row_index, column=8, value=result)
            workbook.save('account.xlsx')
    except Exception as e:
        print(f'Failed to save result: {e}')


def load_accounts_from_excel(file_path):
    """Load account credentials from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        accounts = []
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, 
                                                         min_col=1, max_col=7), start=2):
            email = row[0].value
            password = row[1].value
            recovery_email = row[2].value
            change_recovery = row[3].value
            change_pass = row[4].value
            change_fname = row[5].value
            change_lname = row[6].value
            
            if email and password:
                accounts.append((email, password, recovery_email, change_recovery, 
                               change_pass, change_fname, change_lname, row_index))
        
        return accounts, workbook
    except Exception as e:
        print(f'Error loading Excel file: {e}')
        return [], None


if __name__ == '__main__':
    print('Gmail Master Bot - Standalone Version')
    print('=' * 50)
    
    try:
        accounts, workbook = load_accounts_from_excel('account.xlsx')
        
        if not accounts:
            print('No accounts found in account.xlsx')
            exit(1)
        
        max_threads_input = input('Enter the number of threads you want to use: ').strip()
        max_threads = int(max_threads_input)
        
        if max_threads < 1:
            print('Thread count must be at least 1')
            exit(1)
        
        print(f'Processing {len(accounts)} accounts with {max_threads} threads...')
        
        with ThreadPoolExecutor(max_workers=max_threads) as executor:
            futures = [
                executor.submit(login_to_gmail, email, password, recovery_email, 
                              change_recovery, change_pass, change_fname, change_lname, 
                              row_index, workbook)
                for email, password, recovery_email, change_recovery, change_pass, 
                    change_fname, change_lname, row_index in accounts
            ]
            
            # Wait for all tasks to complete
            for future in futures:
                try:
                    future.result()
                except Exception as e:
                    print(f'Task error: {e}')
        
        print('=' * 50)
        print('All login attempts completed.')
        
    except ValueError:
        print('Invalid input: Please enter a valid number for thread count')
    except Exception as e:
        print(f'Fatal error: {e}')